import json
import logging
from datetime import timedelta
from decimal import Decimal, InvalidOperation
from functools import wraps
from io import BytesIO

from django.contrib import messages
from django.contrib.auth import authenticate, login, logout
from django.contrib.auth.decorators import login_required
from django.contrib.auth.models import User
from django.db import transaction
from django.db.models import Q, Sum
from django.http import HttpResponse, JsonResponse
from django.shortcuts import get_object_or_404, redirect, render
from django.urls import reverse
from django.utils import timezone
from django.views.decorators.http import require_http_methods, require_POST
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Side

from .models import ItemPedido, Pedido, Precio, Producto, SucursalCliente

logger = logging.getLogger(__name__)


def is_admin_user(user):
    return user.is_authenticated and (user.is_staff or user.is_superuser)


def admin_required(view_func):
    @wraps(view_func)
    @login_required
    def wrapped(request, *args, **kwargs):
        if not is_admin_user(request.user):
            messages.error(request, "No tienes permiso para entrar al panel admin.")
            return redirect("pedidos")
        return view_func(request, *args, **kwargs)

    return wrapped


def home(request):
    if not request.user.is_authenticated:
        return redirect("login")
    if is_admin_user(request.user):
        return redirect("admin_dashboard")
    return redirect("pedidos")


@require_http_methods(["GET", "POST"])
def login_view(request):
    """Login compatible con nombres visibles y usernames internos sin espacios."""

    if request.user.is_authenticated:
        return redirect("home")

    if request.method == "POST":
        identificador = request.POST.get("username", "").strip()
        password = request.POST.get("password", "")
        username = identificador

        sucursal = (
            SucursalCliente.objects.select_related("usuario")
            .filter(nombre__iexact=identificador, activa=True)
            .first()
        )
        if sucursal and sucursal.usuario:
            username = sucursal.usuario.username
        else:
            user_match = User.objects.filter(username__iexact=identificador).first()
            if user_match:
                username = user_match.username

        user = authenticate(request, username=username, password=password)
        if user is not None:
            login(request, user)
            return redirect("admin_dashboard" if is_admin_user(user) else "pedidos")

        messages.error(request, "Usuario o contraseña incorrectos.")

    return render(request, "pedidos/login.html")


@require_http_methods(["GET", "POST"])
def logout_view(request):
    logout(request)
    return redirect("login")


def sucursal_para_usuario(user):
    try:
        perfil = user.perfil_pedidos
    except SucursalCliente.DoesNotExist:
        return None
    return perfil if perfil.activa else None


def precio_vigente(producto, sucursal):
    return (
        Precio.objects.filter(
            producto=producto,
            sucursal_cliente=sucursal,
            fecha_vigencia__lte=timezone.localdate(),
        )
        .order_by("-fecha_vigencia")
        .first()
    )


def pedido_pendiente(sucursal, crear=False):
    pedido = (
        Pedido.objects.filter(
            sucursal_cliente=sucursal,
            estado=Pedido.Estado.PENDIENTE,
            eliminado=False,
        )
        .prefetch_related("items__producto")
        .order_by("-fecha_creacion")
        .first()
    )
    if pedido is None and crear:
        pedido = Pedido.objects.create(
            sucursal_cliente=sucursal,
            usuario_nombre=sucursal.nombre,
        )
    return pedido


def decimal_to_str(value, places="0.01"):
    return str(Decimal(value).quantize(Decimal(places)))


def serializar_item(item):
    return {
        "id": item.id,
        "producto_id": item.producto_id,
        "producto": item.producto.nombre,
        "cantidad": decimal_to_str(item.cantidad, "0.001"),
        "precio_unitario": decimal_to_str(item.precio_unitario),
        "subtotal": decimal_to_str(item.subtotal),
    }


def serializar_pedido(pedido):
    if not pedido:
        return {"items": [], "total": "0.00"}
    return {
        "items": [serializar_item(item) for item in pedido.items.all()],
        "total": decimal_to_str(pedido.total),
    }


def parse_json_body(request):
    try:
        return json.loads(request.body.decode("utf-8") or "{}")
    except json.JSONDecodeError:
        return None


def parse_cantidad(value):
    try:
        cantidad = Decimal(str(value)).quantize(Decimal("0.001"))
    except (InvalidOperation, TypeError, ValueError):
        return None
    if cantidad <= 0 or cantidad > Decimal("999.999"):
        return None
    return cantidad


@login_required
def pedidos_view(request):
    if is_admin_user(request.user):
        return redirect("admin_dashboard")

    sucursal = sucursal_para_usuario(request.user)
    if sucursal is None:
        messages.error(request, "Tu usuario no tiene una sucursal o cliente activo.")
        return redirect("login")

    productos = list(Producto.objects.all())
    productos_data = []
    for producto in productos:
        precio = precio_vigente(producto, sucursal)
        productos_data.append(
            {
                "id": producto.id,
                "nombre": producto.nombre,
                "descripcion": producto.descripcion,
                "precio": decimal_to_str(precio.precio_unitario if precio else Decimal("0.00")),
            }
        )

    pedido = pedido_pendiente(sucursal)
    context = {
        "sucursal": sucursal,
        "productos": productos,
        "initial_data": {
            "productos": productos_data,
            "pedido": serializar_pedido(pedido),
        },
    }
    return render(request, "pedidos/pedidos.html", context)


@require_POST
@login_required
def crear_item(request):
    if is_admin_user(request.user):
        return JsonResponse({"success": False, "mensaje": "Admin no puede crear pedidos."}, status=403)

    sucursal = sucursal_para_usuario(request.user)
    if sucursal is None:
        return JsonResponse({"success": False, "mensaje": "Usuario sin sucursal activa."}, status=403)

    payload = parse_json_body(request)
    if payload is None:
        return JsonResponse({"success": False, "mensaje": "JSON inválido."}, status=400)

    producto = get_object_or_404(Producto, pk=payload.get("producto_id"))
    cantidad = parse_cantidad(payload.get("cantidad"))
    if cantidad is None:
        return JsonResponse({"success": False, "mensaje": "Cantidad inválida."}, status=400)

    precio = precio_vigente(producto, sucursal)
    if precio is None:
        return JsonResponse({"success": False, "mensaje": "No hay precio vigente."}, status=400)

    with transaction.atomic():
        pedido = pedido_pendiente(sucursal, crear=True)
        item, created = ItemPedido.objects.select_for_update().get_or_create(
            pedido=pedido,
            producto=producto,
            defaults={
                "cantidad": cantidad,
                "precio_unitario": precio.precio_unitario,
            },
        )
        if not created:
            nueva_cantidad = item.cantidad + cantidad
            if nueva_cantidad > Decimal("999.999"):
                return JsonResponse(
                    {"success": False, "mensaje": "La cantidad máxima es 999.999."},
                    status=400,
                )
            item.cantidad = nueva_cantidad
            item.precio_unitario = precio.precio_unitario
            item.save()
        pedido.recalcular_total()

    return JsonResponse(
        {
            "success": True,
            "mensaje": "Producto agregado.",
            "item_id": item.id,
            "subtotal": decimal_to_str(item.subtotal),
            "total_pedido": decimal_to_str(pedido.total),
            "pedido": serializar_pedido(pedido),
        }
    )


@require_POST
@login_required
def eliminar_item(request):
    sucursal = sucursal_para_usuario(request.user)
    payload = parse_json_body(request)
    if sucursal is None or payload is None:
        return JsonResponse({"success": False, "mensaje": "Solicitud inválida."}, status=400)

    pedido = pedido_pendiente(sucursal)
    item = get_object_or_404(
        ItemPedido,
        pk=payload.get("item_id"),
        pedido=pedido,
        pedido__estado=Pedido.Estado.PENDIENTE,
    )
    item.delete()
    pedido.recalcular_total()
    return JsonResponse(
        {
            "success": True,
            "total_pedido": decimal_to_str(pedido.total),
            "pedido": serializar_pedido(pedido),
        }
    )


@require_POST
@login_required
def limpiar_pedido(request):
    sucursal = sucursal_para_usuario(request.user)
    if sucursal is None:
        return JsonResponse({"success": False, "mensaje": "Usuario sin sucursal activa."}, status=403)
    pedido = pedido_pendiente(sucursal)
    if pedido:
        pedido.items.all().delete()
        pedido.recalcular_total()
    return JsonResponse({"success": True, "pedido": {"items": [], "total": "0.00"}})


@require_POST
@login_required
def confirmar_pedido(request):
    if is_admin_user(request.user):
        return JsonResponse({"success": False, "mensaje": "Admin no puede confirmar pedidos."}, status=403)

    sucursal = sucursal_para_usuario(request.user)
    if sucursal is None:
        return JsonResponse({"success": False, "mensaje": "Usuario sin sucursal activa."}, status=403)

    limite = timezone.now() - timedelta(seconds=60)
    if Pedido.objects.filter(
        sucursal_cliente=sucursal,
        fecha_confirmacion__gte=limite,
        estado__in=[Pedido.Estado.CONFIRMADO, Pedido.Estado.ENVIADO, Pedido.Estado.RECIBIDO],
        eliminado=False,
    ).exists():
        return JsonResponse(
            {"success": False, "mensaje": "Espera un minuto antes de confirmar otro pedido."},
            status=429,
        )

    with transaction.atomic():
        pedido = (
            Pedido.objects.select_for_update()
            .filter(
                sucursal_cliente=sucursal,
                estado=Pedido.Estado.PENDIENTE,
                eliminado=False,
            )
            .prefetch_related("items")
            .order_by("-fecha_creacion")
            .first()
        )
        if pedido is None or not pedido.items.exists():
            return JsonResponse(
                {"success": False, "mensaje": "No hay productos en el pedido."},
                status=400,
            )

        pedido.recalcular_total()
        pedido.estado = Pedido.Estado.CONFIRMADO
        pedido.fecha_confirmacion = timezone.now()
        pedido.save(update_fields=["estado", "fecha_confirmacion"])

    logger.info("Pedido confirmado #%s por %s", pedido.id, sucursal.nombre)
    return JsonResponse(
        {
            "success": True,
            "pedido_id": pedido.id,
            "total": decimal_to_str(pedido.total),
            "mensaje": f"Pedido confirmado #{pedido.id}.",
        }
    )


@admin_required
def admin_dashboard(request):
    pedidos = (
        Pedido.objects.filter(eliminado=False)
        .select_related("sucursal_cliente")
        .prefetch_related("items__producto")
        .order_by("-fecha_creacion")
    )

    estado = request.GET.get("estado", "").strip()
    sucursal_id = request.GET.get("sucursal", "").strip()
    desde = request.GET.get("desde", "").strip()
    hasta = request.GET.get("hasta", "").strip()
    q = request.GET.get("q", "").strip()

    if estado:
        pedidos = pedidos.filter(estado=estado)
    if sucursal_id:
        pedidos = pedidos.filter(sucursal_cliente_id=sucursal_id)
    if desde:
        pedidos = pedidos.filter(fecha_creacion__date__gte=desde)
    if hasta:
        pedidos = pedidos.filter(fecha_creacion__date__lte=hasta)
    if q:
        filtro = Q(sucursal_cliente__nombre__icontains=q) | Q(usuario_nombre__icontains=q)
        if q.isdigit():
            filtro |= Q(id=int(q))
        pedidos = pedidos.filter(filtro)

    pedidos_list = list(pedidos)
    for pedido in pedidos_list:
        pedido.items_json = json.dumps([serializar_item(item) for item in pedido.items.all()])

    hoy = timezone.localdate()
    stats_base = Pedido.objects.filter(eliminado=False)
    stats = {
        "pendientes": stats_base.filter(estado=Pedido.Estado.CONFIRMADO).count(),
        "total_hoy": stats_base.filter(fecha_creacion__date=hoy).aggregate(total=Sum("total"))["total"]
        or Decimal("0.00"),
        "pedidos_hoy": stats_base.filter(fecha_creacion__date=hoy).count(),
    }

    context = {
        "pedidos": pedidos_list,
        "sucursales": SucursalCliente.objects.filter(activa=True),
        "estados": Pedido.Estado.choices,
        "stats": stats,
        "filters": {
            "estado": estado,
            "sucursal": sucursal_id,
            "desde": desde,
            "hasta": hasta,
            "q": q,
        },
    }
    return render(request, "pedidos/admin_dashboard.html", context)


def excel_response_for_pedido(pedido):
    wb = Workbook()
    ws = wb.active
    ws.title = f"Pedido {pedido.id}"
    thin = Side(style="thin", color="D9E2E8")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for row_idx, item in enumerate(pedido.items.select_related("producto").all(), start=1):
        ws.cell(row=row_idx, column=1, value=item.producto.nombre)
        ws.cell(row=row_idx, column=2, value=float(item.cantidad))
        ws.cell(row=row_idx, column=3, value="")
        for col_idx in range(1, 4):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.border = border
            cell.alignment = Alignment(vertical="center")

    ws.column_dimensions["A"].width = 24
    ws.column_dimensions["B"].width = 14
    ws.column_dimensions["C"].width = 8

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    fecha = timezone.localtime(pedido.fecha_creacion).strftime("%Y%m%d")
    filename = f"pedido_{pedido.id}_{fecha}.xlsx"
    response = HttpResponse(
        output.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    return response


@admin_required
def descargar_excel(request, pedido_id):
    pedido = get_object_or_404(
        Pedido.objects.prefetch_related("items__producto"),
        pk=pedido_id,
        eliminado=False,
    )
    logger.info("Admin %s descargó Excel de pedido #%s", request.user.username, pedido.id)
    return excel_response_for_pedido(pedido)


@admin_required
def descargar_y_marcar(request, pedido_id):
    pedido = get_object_or_404(Pedido, pk=pedido_id, eliminado=False)
    if pedido.estado == Pedido.Estado.CONFIRMADO:
        pedido.estado = Pedido.Estado.ENVIADO
        pedido.save(update_fields=["estado"])
        logger.info("Pedido #%s marcado enviado por descarga de %s", pedido.id, request.user.username)
    return excel_response_for_pedido(pedido)


@require_POST
@admin_required
def marcar_enviado(request, pedido_id):
    pedido = get_object_or_404(Pedido, pk=pedido_id, eliminado=False)
    pedido.estado = Pedido.Estado.ENVIADO
    pedido.save(update_fields=["estado"])
    logger.info("Pedido #%s marcado enviado por %s", pedido.id, request.user.username)
    return redirect(f"{reverse('admin_dashboard')}?estado={Pedido.Estado.ENVIADO}")


@require_POST
@admin_required
def eliminar_pedido(request, pedido_id):
    pedido = get_object_or_404(Pedido, pk=pedido_id, eliminado=False)
    pedido.eliminado = True
    pedido.save(update_fields=["eliminado"])
    logger.info("Pedido #%s eliminado suavemente por %s", pedido.id, request.user.username)
    messages.success(request, f"Pedido #{pedido.id} eliminado.")
    return redirect("admin_dashboard")
