import json
from datetime import datetime, time, timedelta
from decimal import Decimal
from io import BytesIO
from pathlib import Path

from django.conf import settings
from django.contrib.auth.models import User
from django.core import mail
from django.core.cache import cache
from django.core.management import call_command
from django.test import TestCase
from django.utils import timezone
from openpyxl import load_workbook

from .models import (
    CONFIGURACION_CACHE_KEY,
    Configuracion,
    ItemPedido,
    LogRecordatorio,
    Pedido,
    Precio,
    Producto,
    SucursalCliente,
)
from .seed import CLIENTES_DEMO, password_for_cliente, seed_demo_data


def abrir_horario_completo():
    """Deja el horario de pedidos abierto todo el día, para que las pruebas que
    no son sobre restricción horaria no dependan de la hora real de ejecución."""

    config = Configuracion.get_solo()
    config.hora_inicio_pedidos = time(0, 0)
    config.hora_fin_pedidos = time(23, 59)
    config.save()
    cache.delete(CONFIGURACION_CACHE_KEY)
    return config


class PedidoFlowTests(TestCase):
    def setUp(self):
        seed_demo_data()
        abrir_horario_completo()

    def test_configuracion_estaticos_mantiene_whitenoise(self):
        self.assertIn("whitenoise.middleware.WhiteNoiseMiddleware", settings.MIDDLEWARE)

    def test_css_movil_no_bloquea_scroll_global(self):
        responsive_css = Path(settings.BASE_DIR, "static", "css", "responsive.css").read_text(
            encoding="utf-8"
        )
        self.assertNotIn("html,\n    body.order-page", responsive_css.replace("\r\n", "\n"))
        self.assertIn("body.order-page {\n        height: 100dvh;", responsive_css.replace("\r\n", "\n"))

    def crear_pedido_confirmado(self, sucursal_nombre, items, fecha_confirmacion=None):
        sucursal = SucursalCliente.objects.get(nombre=sucursal_nombre)
        pedido = Pedido.objects.create(
            sucursal_cliente=sucursal,
            usuario_nombre=sucursal.nombre,
            estado=Pedido.Estado.CONFIRMADO,
            fecha_confirmacion=fecha_confirmacion or timezone.now(),
        )
        for producto_nombre, cantidad in items:
            producto = Producto.objects.get(nombre=producto_nombre)
            ItemPedido.objects.create(
                pedido=pedido,
                producto=producto,
                cantidad=Decimal(str(cantidad)),
                precio_unitario=Decimal("1.00"),
            )
        pedido.recalcular_total()
        return pedido

    def test_login_crear_confirmar_y_excel(self):
        self.assertTrue(self.client.login(username="aguilas", password="Aguilas8445"))
        producto = Producto.objects.get(nombre="LITRO DE BARBACOA")

        response = self.client.post(
            "/api/pedidos/crear-item/",
            data=json.dumps({"producto_id": producto.id, "cantidad": "2.5"}),
            content_type="application/json",
        )
        self.assertEqual(response.status_code, 200)
        data = response.json()
        self.assertTrue(data["success"])
        self.assertEqual(data["total_pedido"], "445.00")

        response = self.client.post("/api/pedidos/confirmar/", content_type="application/json")
        self.assertEqual(response.status_code, 200)
        data = response.json()
        self.assertTrue(data["success"])
        self.assertIn("pedido_folio", data)
        self.assertNotIn("#", data["mensaje"])

        pedido = Pedido.objects.get(id=data["pedido_id"])
        self.assertEqual(pedido.estado, Pedido.Estado.CONFIRMADO)
        self.assertEqual(pedido.total, Decimal("445.00"))

        self.client.logout()
        self.assertTrue(self.client.login(username="juancarlos", password="TocayosMO2026"))
        response = self.client.get(f"/admin/pedidos/{pedido.id}/excel/")
        self.assertEqual(response.status_code, 200)
        workbook = load_workbook(BytesIO(response.content))
        sheet = workbook.active
        self.assertEqual(sheet.title, "PEDIDOS")
        self.assertIn("A1:C1", [str(range_ref) for range_ref in sheet.merged_cells.ranges])
        self.assertEqual(sheet["A1"].value, "AGUILAS")
        self.assertEqual(sheet["A3"].value, "BARBACOA")
        self.assertEqual(sheet["B3"].value, "2.5 KG")
        self.assertEqual(sheet["C2"].number_format, "d-mmm")
        self.assertAlmostEqual(sheet.column_dimensions["A"].width, 15.140625)
        self.assertAlmostEqual(sheet.column_dimensions["B"].width, 10.7109375)
        self.assertAlmostEqual(sheet.column_dimensions["C"].width, 10.140625)
        self.assertEqual(sheet.row_dimensions[1].height, 21.75)
        self.assertEqual(sheet.row_dimensions[2].height, 16.5)
        self.assertEqual(sheet.row_dimensions[3].height, 26.25)
        self.assertEqual(sheet.max_row, 3)
        self.assertEqual(sheet.page_margins.left, 0)
        self.assertEqual(sheet.page_margins.right, 0)
        self.assertEqual(sheet.page_setup.paperSize, 121)
        self.assertEqual(sheet.page_setup.scale, 90)
        self.assertEqual(sheet.page_setup.horizontalDpi, 203)
        self.assertEqual(sheet.print_area, "'PEDIDOS'!$A$1:$C$3")

        print_response = self.client.get(f"/admin/pedidos/{pedido.id}/imprimir/")
        self.assertEqual(print_response.status_code, 200)
        self.assertContains(print_response, "window.print()")
        self.assertContains(print_response, "size: 72mm 73mm;")
        self.assertContains(print_response, 'class="ticket-item-row"', count=1)
        self.assertContains(print_response, "AGUILAS")
        self.assertContains(print_response, "BARBACOA")
        self.assertContains(print_response, "2.5 KG")
        self.assertNotContains(print_response, "$445.00")

        dashboard = self.client.get("/admin/")
        self.assertContains(dashboard, pedido.folio_fecha)
        self.assertContains(dashboard, "data-inline-print")
        self.assertContains(dashboard, f'data-print-template-id="print-pedido-{pedido.id}"')
        self.assertContains(dashboard, 'id="inlinePrintSurface"')
        self.assertContains(dashboard, f'id="print-pedido-{pedido.id}"')
        self.assertNotContains(dashboard, 'target="_blank"')
        self.assertNotContains(dashboard, f"#{pedido.id}")

        embedded_print = self.client.get(f"/admin/pedidos/{pedido.id}/imprimir/?embedded=1")
        self.assertEqual(embedded_print.status_code, 200)
        self.assertNotContains(embedded_print, 'window.addEventListener("load"')

    def test_ticket_imprime_solo_una_fila_por_producto_pedido(self):
        self.assertTrue(self.client.login(username="aguilas", password="Aguilas8445"))
        productos = list(
            Producto.objects.filter(precios__sucursal_cliente__nombre="Aguilas")
            .distinct()
            .order_by("orden", "nombre")[:5]
        )
        self.assertEqual(len(productos), 5)
        for producto in productos:
            response = self.client.post(
                "/api/pedidos/crear-item/",
                data=json.dumps({"producto_id": producto.id, "cantidad": "1"}),
                content_type="application/json",
            )
            self.assertEqual(response.status_code, 200)

        response = self.client.post("/api/pedidos/confirmar/", content_type="application/json")
        pedido_id = response.json()["pedido_id"]

        self.client.logout()
        self.assertTrue(self.client.login(username="juancarlos", password="TocayosMO2026"))
        response = self.client.get(f"/admin/pedidos/{pedido_id}/excel/")
        workbook = load_workbook(BytesIO(response.content))
        self.assertEqual(workbook.active.print_area, "'PEDIDOS'!$A$1:$C$7")

        print_response = self.client.get(f"/admin/pedidos/{pedido_id}/imprimir/")
        self.assertContains(print_response, 'class="ticket-item-row"', count=5)

    def test_cliente_mayorista_usa_precio_dos_pesos(self):
        self.assertTrue(self.client.login(username="brot_nueva_galicia", password="Brot Nueva Galicia0846"))
        producto = Producto.objects.get(nombre="TORTILLA ESPECIAL")
        response = self.client.post(
            "/api/pedidos/crear-item/",
            data=json.dumps({"producto_id": producto.id, "cantidad": "3"}),
            content_type="application/json",
        )
        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.json()["total_pedido"], "79.50")

    def test_chile_guero_se_cobra_por_kilo_promedio_de_treinta_piezas(self):
        self.assertTrue(self.client.login(username="aguilas", password="Aguilas8445"))
        producto = Producto.objects.get(nombre="CHILE GüERO")
        response = self.client.post(
            "/api/pedidos/crear-item/",
            data=json.dumps({"producto_id": producto.id, "cantidad": "30"}),
            content_type="application/json",
        )

        self.assertEqual(response.status_code, 200)
        data = response.json()
        self.assertEqual(data["total_pedido"], "64.00")
        self.assertEqual(data["pedido"]["items"][0]["unidad"], "PZA")

    def test_ticket_mayoreo_marca_productos_con_sufijo_m(self):
        self.assertTrue(self.client.login(username="brot_nueva_galicia", password="Brot Nueva Galicia0846"))
        producto = Producto.objects.get(nombre="LITRO DE BARBACOA")
        response = self.client.post(
            "/api/pedidos/crear-item/",
            data=json.dumps({"producto_id": producto.id, "cantidad": "1"}),
            content_type="application/json",
        )
        self.assertEqual(response.status_code, 200)
        response = self.client.post("/api/pedidos/confirmar/", content_type="application/json")
        pedido_id = response.json()["pedido_id"]

        self.client.logout()
        self.assertTrue(self.client.login(username="juancarlos", password="TocayosMO2026"))
        response = self.client.get(f"/admin/pedidos/{pedido_id}/excel/")
        workbook = load_workbook(BytesIO(response.content))
        self.assertEqual(workbook.active["A3"].value, "BARBACOA .M")

    def test_promo_martes_aguilas_agrega_cinco_por_cada_veinte_en_ticket(self):
        self.assertTrue(self.client.login(username="aguilas", password="Aguilas8445"))
        producto = Producto.objects.get(nombre="LITRO DE BARBACOA")
        response = self.client.post(
            "/api/pedidos/crear-item/",
            data=json.dumps({"producto_id": producto.id, "cantidad": "20"}),
            content_type="application/json",
        )
        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.json()["total_pedido"], "3560.00")
        response = self.client.post("/api/pedidos/confirmar/", content_type="application/json")
        pedido_id = response.json()["pedido_id"]

        pedido = Pedido.objects.get(id=pedido_id)
        martes = timezone.make_aware(datetime(2026, 7, 14, 10, 0))
        pedido.fecha_confirmacion = martes
        pedido.save(update_fields=["fecha_confirmacion"])

        self.client.logout()
        self.assertTrue(self.client.login(username="juancarlos", password="TocayosMO2026"))
        response = self.client.get(f"/admin/pedidos/{pedido_id}/excel/")
        workbook = load_workbook(BytesIO(response.content))
        sheet = workbook.active
        self.assertEqual(sheet["B3"].value, "25 KG")
        self.assertEqual(pedido.total, Decimal("3560.00"))

    def test_agregar_producto_existente_reemplaza_cantidad(self):
        self.assertTrue(self.client.login(username="aguilas", password="Aguilas8445"))
        producto = Producto.objects.get(nombre="LITRO DE BARBACOA")

        first = self.client.post(
            "/api/pedidos/crear-item/",
            data=json.dumps({"producto_id": producto.id, "cantidad": "5"}),
            content_type="application/json",
        )
        second = self.client.post(
            "/api/pedidos/crear-item/",
            data=json.dumps({"producto_id": producto.id, "cantidad": "3"}),
            content_type="application/json",
        )

        self.assertEqual(first.status_code, 200)
        self.assertEqual(second.status_code, 200)
        data = second.json()
        self.assertEqual(data["total_pedido"], "534.00")
        self.assertEqual(len(data["pedido"]["items"]), 1)
        self.assertEqual(data["pedido"]["items"][0]["cantidad"], "3.000")

    def test_primer_item_se_muestra_si_habia_pedido_pendiente_vacio(self):
        self.assertTrue(self.client.login(username="aguilas", password="Aguilas8445"))
        sucursal = SucursalCliente.objects.get(nombre="Aguilas")
        Pedido.objects.create(sucursal_cliente=sucursal, usuario_nombre=sucursal.nombre)
        producto = Producto.objects.get(nombre="SALSA DE AGUACATE")

        response = self.client.post(
            "/api/pedidos/crear-item/",
            data=json.dumps({"producto_id": producto.id, "cantidad": "1"}),
            content_type="application/json",
        )

        self.assertEqual(response.status_code, 200)
        data = response.json()
        self.assertEqual(data["total_pedido"], "60.00")
        self.assertEqual(len(data["pedido"]["items"]), 1)
        self.assertEqual(data["pedido"]["items"][0]["producto"], "SALSA DE AGUACATE")
        self.assertEqual(data["pedido"]["items"][0]["cantidad"], "1.000")

    def test_pantalla_pedidos_no_muestra_precios_unitarios(self):
        self.assertTrue(self.client.login(username="aguilas", password="Aguilas8445"))
        response = self.client.get("/pedidos/")

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, 'href="/pedidos/historial/"')
        self.assertContains(response, ">Historial</a>")
        self.assertNotContains(response, '<span class="brand-title">Pedidos</span>')
        self.assertNotContains(response, "$178.00")
        self.assertNotContains(response, "precio_unitario")
        self.assertNotContains(response, "scheduleStatus")
        self.assertNotContains(response, "Total tentativo")

    def test_usuario_ve_historial_propio_e_imprime_recibo_adaptable(self):
        fecha_reciente = timezone.make_aware(datetime(2026, 7, 17, 12, 0))
        fecha_anterior = timezone.make_aware(datetime(2026, 7, 16, 12, 0))
        fecha_otro = timezone.make_aware(datetime(2026, 7, 15, 12, 0))
        reciente = self.crear_pedido_confirmado(
            "Aguilas",
            [
                ("LITRO DE BARBACOA", "2"),
                ("TORTILLA ESPECIAL", "3"),
            ],
            fecha_reciente,
        )
        anterior = self.crear_pedido_confirmado(
            "Aguilas",
            [("AGUA JAMAICA LT", "4")],
            fecha_anterior,
        )
        otro = self.crear_pedido_confirmado(
            "Fortin",
            [("LITRO DE BARBACOA", "5")],
            fecha_otro,
        )

        self.assertTrue(self.client.login(username="aguilas", password="Aguilas8445"))
        response = self.client.get("/pedidos/historial/")
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Historial")
        self.assertContains(response, f"Pedido {reciente.folio_fecha}")
        self.assertContains(response, f"Pedido {anterior.folio_fecha}")
        self.assertContains(response, str(reciente.codigo_publico))
        self.assertNotContains(response, f"Pedido {otro.folio_fecha}")
        self.assertNotContains(response, f"Pedido #{reciente.id}")
        html = response.content.decode()
        self.assertLess(
            html.index(f"Pedido {reciente.folio_fecha}"),
            html.index(f"Pedido {anterior.folio_fecha}"),
        )
        self.assertContains(response, 'data-print-size="auto"')
        self.assertContains(
            response,
            f'data-print-template-id="print-history-pedido-{reciente.codigo_publico}"',
        )
        self.assertContains(response, "Total provisional")
        self.assertNotContains(response, "$1.00")
        self.assertNotContains(response, "precio_unitario")

        print_response = self.client.get(f"/pedidos/historial/{reciente.codigo_publico}/imprimir/")
        self.assertEqual(print_response.status_code, 200)
        self.assertContains(print_response, "size: auto;")
        self.assertContains(print_response, "window.print()")
        self.assertContains(print_response, "Aguilas")
        self.assertContains(print_response, reciente.folio_fecha)
        self.assertContains(print_response, str(reciente.codigo_publico).split("-")[0].upper())
        self.assertContains(print_response, "LITRO DE BARBACOA")
        self.assertContains(print_response, "2 KG")
        self.assertContains(print_response, "Total provisional")
        self.assertNotContains(print_response, "$1.00")
        self.assertNotContains(print_response, "precio_unitario")
        self.assertNotContains(print_response, f"#{reciente.id}")

        embedded = self.client.get(f"/pedidos/historial/{reciente.codigo_publico}/imprimir/?embedded=1")
        self.assertEqual(embedded.status_code, 200)
        self.assertNotContains(embedded, 'window.addEventListener("load"')

        forbidden = self.client.get(f"/pedidos/historial/{otro.codigo_publico}/imprimir/")
        self.assertEqual(forbidden.status_code, 404)

    def test_login_muestra_horario_de_pedidos(self):
        response = self.client.get("/login/")
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Pedidos abiertos")
        self.assertContains(response, "data-password-toggle")
        self.assertContains(response, 'aria-label="Mostrar contraseña"')
        self.assertContains(response, "Sistema privado de uso exclusivo")
        self.assertContains(response, 'href="/privacidad/"')

    def test_aviso_privacidad_publico_y_link_interno(self):
        response = self.client.get("/privacidad/")
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Aviso de privacidad")
        self.assertContains(response, "Cookies técnicas")
        self.assertContains(response, "tocayos.tacos@gmail.com")

        self.assertTrue(self.client.login(username="aguilas", password="Aguilas8445"))
        response = self.client.get("/pedidos/")
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, 'class="privacy-toplink"')
        self.assertContains(response, 'href="/privacidad/"')

    def test_usuario_no_admin_no_puede_ver_dashboard(self):
        self.assertTrue(self.client.login(username="fortin", password="Fortin9481"))
        response = self.client.get("/admin/")
        self.assertEqual(response.status_code, 302)
        self.assertEqual(response.url, "/pedidos/")

        response = self.client.get("/admin/configuracion/")
        self.assertEqual(response.status_code, 302)
        self.assertEqual(response.url, "/pedidos/")

    def test_usuario_impresion_solo_ve_dashboard_e_imprime(self):
        self.assertTrue(self.client.login(username="aguilas", password="Aguilas8445"))
        producto = Producto.objects.get(nombre="LITRO DE BARBACOA")
        self.client.post(
            "/api/pedidos/crear-item/",
            data=json.dumps({"producto_id": producto.id, "cantidad": "1"}),
            content_type="application/json",
        )
        response = self.client.post("/api/pedidos/confirmar/", content_type="application/json")
        pedido_id = response.json()["pedido_id"]
        self.client.logout()

        self.assertTrue(self.client.login(username="juanmanuel", password="imprimir"))
        dashboard = self.client.get("/admin/")
        self.assertEqual(dashboard.status_code, 200)
        self.assertContains(dashboard, "Ver detalle")
        self.assertContains(dashboard, "Imprimir")
        self.assertContains(dashboard, "Aguas")
        self.assertContains(dashboard, "data-inline-print")
        self.assertContains(dashboard, 'data-print-template-id="print-aguas"')
        self.assertContains(dashboard, f'id="print-pedido-{pedido_id}"')
        html = dashboard.content.decode()
        self.assertNotIn('target="_blank"', html)
        self.assertNotIn("admin/configuracion", html)
        self.assertNotIn("admin/datos", html)
        self.assertNotIn(">Excel</a>", html)
        self.assertNotIn("marcar-enviado", html)
        self.assertNotIn("eliminar/", html)

        print_response = self.client.get(f"/admin/pedidos/{pedido_id}/imprimir/")
        self.assertEqual(print_response.status_code, 200)
        self.assertContains(print_response, "window.print()")
        aguas_response = self.client.get("/admin/aguas/imprimir/")
        self.assertEqual(aguas_response.status_code, 200)
        self.assertContains(aguas_response, "size: 72mm 72mm;")
        aguas_embedded = self.client.get("/admin/aguas/imprimir/?embedded=1")
        self.assertEqual(aguas_embedded.status_code, 200)
        self.assertNotContains(aguas_embedded, 'window.addEventListener("load"')

        response = self.client.get("/admin/configuracion/")
        self.assertEqual(response.status_code, 302)
        response = self.client.get("/admin/datos/")
        self.assertEqual(response.status_code, 302)
        response = self.client.get(f"/admin/pedidos/{pedido_id}/excel/")
        self.assertEqual(response.status_code, 302)
        response = self.client.post(f"/admin/pedidos/{pedido_id}/eliminar/")
        self.assertEqual(response.status_code, 302)
        self.assertFalse(Pedido.objects.get(id=pedido_id).eliminado)

    def test_admin_configura_ticket_precio_y_password(self):
        self.assertTrue(self.client.login(username="juancarlos", password="TocayosMO2026"))
        producto = Producto.objects.get(nombre="LITRO DE BARBACOA")
        sucursal = SucursalCliente.objects.get(nombre="Aguilas")

        response = self.client.get("/admin/configuracion/")
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Base de datos")
        self.assertNotContains(response, "pbkdf2_")

        response = self.client.post(
            "/admin/configuracion/",
            data={
                "action": "actualizar_productos",
                f"producto_{producto.id}_present": "1",
                f"producto_{producto.id}_nombre": "LITRO DE BARBACOA",
                f"producto_{producto.id}_ticket": "BARBA",
                f"producto_{producto.id}_unidad": producto.unidad_medida,
                f"producto_{producto.id}_unidad_abreviatura": producto.unidad_abreviatura,
                f"producto_{producto.id}_cantidad_por_precio": str(producto.cantidad_por_precio),
                f"producto_{producto.id}_orden": str(producto.orden),
                f"producto_{producto.id}_activo": "on",
            },
        )
        self.assertRedirects(response, "/admin/configuracion/")

        response = self.client.post(
            "/admin/configuracion/",
            data={
                "action": "actualizar_precios",
                f"precio_{producto.id}_{sucursal.id}": "12.50",
                f"precio_ticket_{producto.id}_{sucursal.id}": "BARBA",
            },
        )
        self.assertRedirects(response, "/admin/configuracion/")
        precio = Precio.objects.get(
            producto=producto,
            sucursal_cliente=sucursal,
            fecha_vigencia__isnull=False,
        )
        self.assertEqual(precio.precio_unitario, Decimal("12.50"))

        response = self.client.post(
            "/admin/configuracion/",
            data={
                "action": "actualizar_sucursales",
                f"sucursal_{sucursal.id}_present": "1",
                f"sucursal_{sucursal.id}_nombre": sucursal.nombre,
                f"sucursal_{sucursal.id}_tipo": sucursal.tipo,
                f"sucursal_{sucursal.id}_username": sucursal.usuario.username,
                f"sucursal_{sucursal.id}_email": sucursal.email,
                f"sucursal_{sucursal.id}_password": "NuevaClave123",
                f"sucursal_{sucursal.id}_activa": "on",
            },
        )
        self.assertRedirects(response, "/admin/configuracion/")

        self.client.logout()
        self.assertFalse(self.client.login(username="aguilas", password="Aguilas8445"))
        self.assertTrue(self.client.login(username="aguilas", password="NuevaClave123"))
        response = self.client.post(
            "/api/pedidos/crear-item/",
            data=json.dumps({"producto_id": producto.id, "cantidad": "2"}),
            content_type="application/json",
        )
        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.json()["total_pedido"], "25.00")
        response = self.client.post("/api/pedidos/confirmar/", content_type="application/json")
        pedido_id = response.json()["pedido_id"]

        self.client.logout()
        self.assertTrue(self.client.login(username="juancarlos", password="TocayosMO2026"))
        response = self.client.get(f"/admin/pedidos/{pedido_id}/excel/")
        workbook = load_workbook(BytesIO(response.content))
        self.assertEqual(workbook.active["A3"].value, "BARBA")

    def test_admin_puede_crear_producto_y_usuario(self):
        self.assertTrue(self.client.login(username="juancarlos", password="TocayosMO2026"))
        response = self.client.post(
            "/admin/configuracion/",
            data={
                "action": "crear_producto",
                "nuevo_nombre": "Producto Prueba",
                "nuevo_ticket": "PRUEBA",
                "nuevo_unidad": "LITRO (LT)",
                "nuevo_unidad_abreviatura": "LT",
                "nuevo_cantidad_por_precio": "1.000",
                "nuevo_orden": "7",
            },
        )
        self.assertRedirects(response, "/admin/configuracion/")
        self.assertTrue(
            Producto.objects.filter(nombre="Producto Prueba", nombre_ticket="PRUEBA", activo=True).exists()
        )

        response = self.client.post(
            "/admin/configuracion/",
            data={
                "action": "crear_sucursal",
                "nuevo_sucursal_nombre": "Sucursal Prueba",
                "nuevo_sucursal_tipo": SucursalCliente.Tipo.SUCURSAL,
                "nuevo_sucursal_username": "sucursal_prueba",
                "nuevo_sucursal_password": "ClavePrueba123",
            },
        )
        self.assertRedirects(response, "/admin/configuracion/")
        created_user = User.objects.get(username="sucursal_prueba")
        self.assertTrue(created_user.check_password("ClavePrueba123"))
        self.assertTrue(
            SucursalCliente.objects.filter(nombre="Sucursal Prueba", usuario=created_user).exists()
        )

    def test_admin_imprime_aguas_de_pedidos_del_dia_anterior(self):
        ayer = timezone.localdate() - timedelta(days=1)
        fecha_ayer = timezone.make_aware(datetime.combine(ayer, time(11, 0)))
        self.crear_pedido_confirmado(
            "Estancia",
            [
                ("AGUA HORCHATA BLANCA 1/2", "2"),
                ("AGUA HORCHATA ROSA LT", "3"),
            ],
            fecha_ayer,
        )
        self.crear_pedido_confirmado(
            "Aguilas",
            [
                ("AGUA HORCHATA BLANCA 1/2", "5"),
                ("AGUA JAMAICA LT", "1"),
            ],
            fecha_ayer,
        )
        self.crear_pedido_confirmado(
            "Fortin",
            [("LITRO DE BARBACOA", "4")],
            fecha_ayer,
        )

        self.assertTrue(self.client.login(username="juancarlos", password="TocayosMO2026"))
        dashboard = self.client.get("/admin/")
        self.assertContains(dashboard, "Aguas")
        self.assertContains(dashboard, "data-inline-print")
        self.assertContains(dashboard, 'id="print-aguas"')
        self.assertNotContains(dashboard, 'target="_blank"')
        response = self.client.get("/admin/aguas/imprimir/")
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "window.print()")
        self.assertContains(response, "size: 72mm 72mm;")
        self.assertContains(response, "<td>1/B</td>", html=True)
        self.assertContains(response, "<td>2</td>", html=True)
        self.assertContains(response, "<td>5</td>", html=True)
        self.assertContains(response, "<td>/</td>", html=True)
        self.assertContains(response, "<td>LR</td>", html=True)
        self.assertContains(response, "<td>3</td>", html=True)
        self.assertContains(response, "<td>LJ</td>", html=True)
        self.assertContains(response, "<td>1</td>", html=True)

    def test_admin_datos_muestra_promedios_y_prediccion(self):
        lunes = timezone.make_aware(datetime(2026, 7, 13, 10, 0))
        self.crear_pedido_confirmado(
            "Aguilas",
            [
                ("LITRO DE BARBACOA", "2"),
                ("AGUA JAMAICA LT", "4"),
            ],
            lunes,
        )
        self.crear_pedido_confirmado(
            "Aguilas",
            [("AGUA JAMAICA LT", "2")],
            lunes + timedelta(minutes=30),
        )

        sucursal = SucursalCliente.objects.get(nombre="Aguilas")
        self.assertTrue(self.client.login(username="juancarlos", password="TocayosMO2026"))
        response = self.client.get(f"/admin/datos/?sucursal={sucursal.id}&dia=1")
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Datos")
        self.assertContains(response, "Promedio de Lunes")
        self.assertContains(response, "AGUA JAMAICA LT")
        self.assertContains(response, "Ticket promedio")
        self.assertContains(response, "data-inline-print")
        self.assertContains(response, 'data-print-template-id="print-aguas"')
        self.assertContains(response, 'id="print-aguas"')
        self.assertNotContains(response, 'target="_blank"')

    def test_seed_crea_usuario_debug_admin(self):
        user = User.objects.get(username="ivanprueba")
        self.assertTrue(user.is_staff)
        self.assertTrue(user.is_superuser)
        self.assertTrue(user.check_password("prueba8989"))
        self.assertTrue(self.client.login(username="ivanprueba", password="prueba8989"))
        response = self.client.get("/admin/configuracion/")
        self.assertEqual(response.status_code, 200)

    def test_producto_inactivo_solo_permanece_en_pedido_pendiente(self):
        producto = Producto.objects.get(nombre="LITRO DE BARBACOA")
        self.assertTrue(self.client.login(username="aguilas", password="Aguilas8445"))
        response = self.client.post(
            "/api/pedidos/crear-item/",
            data=json.dumps({"producto_id": producto.id, "cantidad": "1"}),
            content_type="application/json",
        )
        self.assertEqual(response.status_code, 200)

        self.client.logout()
        self.assertTrue(self.client.login(username="juancarlos", password="TocayosMO2026"))
        response = self.client.post(
            "/admin/configuracion/",
            data={
                "action": "actualizar_productos",
                f"producto_{producto.id}_present": "1",
                f"producto_{producto.id}_nombre": producto.nombre,
                f"producto_{producto.id}_ticket": producto.nombre_ticket,
                f"producto_{producto.id}_unidad": producto.unidad_medida,
                f"producto_{producto.id}_unidad_abreviatura": producto.unidad_abreviatura,
                f"producto_{producto.id}_cantidad_por_precio": str(producto.cantidad_por_precio),
                f"producto_{producto.id}_orden": str(producto.orden),
            },
        )
        self.assertRedirects(response, "/admin/configuracion/")

        self.client.logout()
        self.assertTrue(self.client.login(username="aguilas", password="Aguilas8445"))
        response = self.client.get("/pedidos/")
        self.assertContains(response, f'data-product-id="{producto.id}"')

        self.client.logout()
        self.assertTrue(self.client.login(username="fortin", password="Fortin9481"))
        response = self.client.get("/pedidos/")
        self.assertNotContains(response, f'data-product-id="{producto.id}"')

    def test_seed_crea_seis_clientes_demo(self):
        self.assertEqual(SucursalCliente.objects.count(), 6)
        for nombre, _, _ in CLIENTES_DEMO:
            sucursal = SucursalCliente.objects.get(nombre=nombre)
            self.assertTrue(sucursal.usuario.check_password(password_for_cliente(nombre)))


class RestriccionHorariaTests(TestCase):
    def setUp(self):
        seed_demo_data()

    def _ventana_fuera_de_ahora(self):
        """Regresa (inicio, fin) que garantizadamente NO incluye la hora actual."""

        ahora = timezone.localtime().time()
        if ahora < time(12, 0):
            return time(20, 0), time(23, 0)
        return time(0, 0), time(1, 0)

    def test_confirmar_pedido_fuera_de_horario_retorna_400(self):
        inicio, fin = self._ventana_fuera_de_ahora()
        config = Configuracion.get_solo()
        config.hora_inicio_pedidos = inicio
        config.hora_fin_pedidos = fin
        config.save()
        cache.delete(CONFIGURACION_CACHE_KEY)

        self.assertTrue(self.client.login(username="aguilas", password="Aguilas8445"))
        producto = Producto.objects.get(nombre="LITRO DE BARBACOA")
        self.client.post(
            "/api/pedidos/crear-item/",
            data=json.dumps({"producto_id": producto.id, "cantidad": "1"}),
            content_type="application/json",
        )

        response = self.client.post("/api/pedidos/confirmar/", content_type="application/json")
        self.assertEqual(response.status_code, 400)
        data = response.json()
        self.assertFalse(data["success"])
        self.assertIn("Pedidos cerrados", data["mensaje"])

        pedido = Pedido.objects.get(sucursal_cliente__nombre="Aguilas")
        self.assertEqual(pedido.estado, Pedido.Estado.PENDIENTE)

    def test_login_muestra_pedidos_cerrados_fuera_de_horario(self):
        inicio, fin = self._ventana_fuera_de_ahora()
        config = Configuracion.get_solo()
        config.hora_inicio_pedidos = inicio
        config.hora_fin_pedidos = fin
        config.save()
        cache.delete(CONFIGURACION_CACHE_KEY)

        response = self.client.get("/login/")
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Pedidos cerrados")

    def test_confirmar_pedido_dentro_de_horario_permite(self):
        abrir_horario_completo()
        self.assertTrue(self.client.login(username="aguilas", password="Aguilas8445"))
        producto = Producto.objects.get(nombre="LITRO DE BARBACOA")
        self.client.post(
            "/api/pedidos/crear-item/",
            data=json.dumps({"producto_id": producto.id, "cantidad": "1"}),
            content_type="application/json",
        )
        response = self.client.post("/api/pedidos/confirmar/", content_type="application/json")
        self.assertEqual(response.status_code, 200)
        self.assertTrue(response.json()["success"])

    def test_endpoint_info_horarios_no_requiere_login(self):
        abrir_horario_completo()
        response = self.client.get("/api/horarios/")
        self.assertEqual(response.status_code, 200)
        data = response.json()
        self.assertTrue(data["dentro_horario"])
        for key in ("hora_inicio", "hora_fin", "hora_actual", "dentro_horario", "mensaje"):
            self.assertIn(key, data)

    def test_admin_actualiza_horarios_desde_panel_configuracion(self):
        self.assertTrue(self.client.login(username="juancarlos", password="TocayosMO2026"))
        response = self.client.post(
            "/admin/configuracion/",
            data={
                "action": "actualizar_configuracion",
                "hora_inicio_pedidos": "09:00",
                "hora_fin_pedidos": "18:00",
                "hora_envio_recordatorio": "15:30",
                "dias_recordatorio": ["1", "2", "3", "4", "5"],
                "email_remitente": "Los Tocayos <tocayos.tacos@gmail.com>",
                "recordatorios_habilitados": "on",
            },
        )
        self.assertRedirects(response, "/admin/configuracion/")

        config = Configuracion.get_solo()
        self.assertEqual(config.hora_inicio_pedidos, time(9, 0))
        self.assertEqual(config.hora_fin_pedidos, time(18, 0))
        self.assertEqual(config.hora_envio_recordatorio, time(15, 30))
        self.assertEqual(config.dias_recordatorio_lista(), [1, 2, 3, 4, 5])
        self.assertTrue(config.recordatorios_habilitados)
        self.assertEqual(config.actualizado_por, "juancarlos")


class EnviarRecordatoriosCommandTests(TestCase):
    def setUp(self):
        seed_demo_data()
        self.sucursal = SucursalCliente.objects.get(nombre="Aguilas")
        self.sucursal.email = "aguilas@example.com"
        self.sucursal.save()

    def test_modo_test_no_envia_correos_reales(self):
        call_command("enviar_recordatorios", test=True, fuerza=True)
        self.assertEqual(len(mail.outbox), 0)

    def test_fuerza_envia_y_registra_log(self):
        call_command("enviar_recordatorios", fuerza=True)
        self.assertEqual(len(mail.outbox), 1)
        self.assertIn(self.sucursal.email, mail.outbox[0].to)
        self.assertIn("tocayos.tacos@gmail.com", mail.outbox[0].from_email)
        self.assertIn("tocayos.tacos@gmail.com", mail.outbox[0].body)
        self.assertTrue(
            LogRecordatorio.objects.filter(
                sucursal_cliente=self.sucursal, estado=LogRecordatorio.Estado.ENVIADO
            ).exists()
        )

    def test_sucursal_sin_correo_se_marca_saltada(self):
        otra = SucursalCliente.objects.get(nombre="Fortin")
        self.assertEqual(otra.email, "")
        call_command("enviar_recordatorios", fuerza=True)
        self.assertTrue(
            LogRecordatorio.objects.filter(
                sucursal_cliente=otra, estado=LogRecordatorio.Estado.SALTADO
            ).exists()
        )

    def test_sin_fuerza_respeta_dias_configurados(self):
        config = Configuracion.get_solo()
        hoy_iso = timezone.localdate().isoweekday()
        # Configura un único día distinto al de hoy para forzar el "saltado".
        otro_dia = 1 if hoy_iso != 1 else 2
        config.dias_recordatorio = str(otro_dia)
        config.save()

        call_command("enviar_recordatorios")
        self.assertEqual(len(mail.outbox), 0)
        self.assertFalse(LogRecordatorio.objects.exists())

    def test_recordatorios_deshabilitados_sin_fuerza_no_envia(self):
        config = Configuracion.get_solo()
        config.recordatorios_habilitados = False
        config.save()

        call_command("enviar_recordatorios")
        self.assertEqual(len(mail.outbox), 0)
