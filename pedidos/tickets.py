from decimal import Decimal
from io import BytesIO

from django.utils import timezone
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, Side
from openpyxl.worksheet.page import PageMargins

from .models import Precio

TICKET_COLUMN_WIDTHS = {
    "A": 15.140625,
    "B": 10.7109375,
    "C": 10.140625,
}
TICKET_ROW_HEIGHTS = {
    1: 21.75,
    2: 16.5,
    7: 23.25,
}
TICKET_DEFAULT_ITEM_ROW_HEIGHT = 26.25
TICKET_HEADER_HEIGHT_MM = 7.67
TICKET_DATE_HEIGHT_MM = 5.82
TICKET_ITEM_HEIGHT_MM = 9.26
TICKET_SHORT_ITEM_HEIGHT_MM = 8.20
TICKET_PRINT_SAFETY_HEIGHT_MM = 6.0
TICKET_WIDTH_MM = 58
TICKET_COLUMN_WIDTHS_MM = {
    "product": 26.25,
    "quantity": 14.85,
    "blank": 16.90,
}
SPANISH_MONTH_ABBR = (
    "ene",
    "feb",
    "mar",
    "abr",
    "may",
    "jun",
    "jul",
    "ago",
    "sep",
    "oct",
    "nov",
    "dic",
)


def ticket_date(pedido):
    base_date = pedido.fecha_confirmacion or pedido.fecha_creacion
    return timezone.localtime(base_date).date()


def ticket_date_display(pedido):
    local_date = ticket_date(pedido)
    return f"{local_date.day}-{SPANISH_MONTH_ABBR[local_date.month - 1]}"


def format_ticket_quantity(value):
    decimal_value = Decimal(value).quantize(Decimal("0.001")).normalize()
    if decimal_value == decimal_value.to_integral():
        return str(decimal_value.quantize(Decimal("1")))
    return format(decimal_value, "f").rstrip("0").rstrip(".")


def ticket_label_for_item(item, fecha):
    precio = (
        Precio.objects.filter(
            producto=item.producto,
            sucursal_cliente=item.pedido.sucursal_cliente,
            fecha_vigencia__lte=fecha,
        )
        .order_by("-fecha_vigencia")
        .first()
    )
    if precio is not None:
        return precio.etiqueta_ticket.upper()
    return item.producto.etiqueta_ticket.upper()


def format_ticket_quantity_with_unit(item, fecha):
    quantity = format_ticket_quantity(item.cantidad_con_promocion(fecha))
    return f"{quantity} {item.producto.unidad_corta}".strip()


def ticket_title(pedido):
    return pedido.sucursal_cliente.nombre.upper()


def ticket_items(pedido):
    fecha = ticket_date(pedido)
    return [
        {
            "producto": ticket_label_for_item(item, fecha),
            "cantidad": format_ticket_quantity_with_unit(item, fecha),
        }
        for item in pedido.items.select_related("producto").all()
    ]


def ticket_context(pedido):
    rows = []
    for index, row in enumerate(ticket_items(pedido), start=3):
        rows.append(
            {
                **row,
                "row_number": index,
                "height_mm": (
                    TICKET_SHORT_ITEM_HEIGHT_MM
                    if index == 7
                    else TICKET_ITEM_HEIGHT_MM
                ),
            }
        )
    ticket_height_mm = (
        TICKET_HEADER_HEIGHT_MM
        + TICKET_DATE_HEIGHT_MM
        + sum(row["height_mm"] for row in rows)
        + TICKET_PRINT_SAFETY_HEIGHT_MM
    )
    return {
        "pedido": pedido,
        "title": ticket_title(pedido),
        "date": ticket_date(pedido),
        "date_display": ticket_date_display(pedido),
        "rows": rows,
        "ticket_width_mm": TICKET_WIDTH_MM,
        "ticket_height_mm": round(ticket_height_mm, 2),
        "column_widths_mm": TICKET_COLUMN_WIDTHS_MM,
        "header_height_mm": TICKET_HEADER_HEIGHT_MM,
        "date_height_mm": TICKET_DATE_HEIGHT_MM,
    }


def build_ticket_workbook(pedido):
    wb = Workbook()
    ws = wb.active
    ws.title = "PEDIDOS"
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A2"

    rows = ticket_items(pedido)
    last_row = len(rows) + 2

    ws.merge_cells("A1:C1")
    ws["A1"] = ticket_title(pedido)
    ws["C2"] = ticket_date(pedido)
    ws["C2"].number_format = "d-mmm"

    for column, width in TICKET_COLUMN_WIDTHS.items():
        ws.column_dimensions[column].width = width

    for row_number in range(1, last_row + 1):
        ws.row_dimensions[row_number].height = TICKET_ROW_HEIGHTS.get(
            row_number,
            TICKET_DEFAULT_ITEM_ROW_HEIGHT,
        )

    thin_black = Side(style="thin", color="000000")
    table_border = Border(
        left=thin_black,
        right=thin_black,
        top=thin_black,
        bottom=thin_black,
    )
    title_font = Font(name="Calibri", size=16, bold=True)
    product_font = Font(name="Calibri", size=10, bold=True)
    quantity_font = Font(name="Calibri", size=9, bold=True)

    ws["A1"].font = title_font
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws["C2"].font = Font(name="Calibri", size=11, bold=True)
    ws["C2"].alignment = Alignment(horizontal="center", vertical="center")
    ws["C2"].border = table_border

    for offset, row in enumerate(rows, start=3):
        product_cell = ws.cell(row=offset, column=1, value=row["producto"])
        quantity_cell = ws.cell(row=offset, column=2, value=row["cantidad"])
        blank_cell = ws.cell(row=offset, column=3, value="")

        product_cell.font = product_font
        product_cell.alignment = Alignment(vertical="center")
        quantity_cell.font = quantity_font
        quantity_cell.alignment = Alignment(horizontal="center", vertical="center")
        blank_cell.alignment = Alignment(vertical="center")

        for cell in (product_cell, quantity_cell, blank_cell):
            cell.border = table_border

    ws.print_area = f"A1:C{last_row}"
    ws.page_margins = PageMargins(
        left=0,
        right=0,
        top=0,
        bottom=0,
        header=0.31496062992125984,
        footer=0.31496062992125984,
    )
    ws.page_setup.paperSize = "121"
    ws.page_setup.scale = 90
    ws.page_setup.orientation = "portrait"
    ws.page_setup.horizontalDpi = 203
    ws.page_setup.verticalDpi = 203

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output
